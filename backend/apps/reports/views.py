from datetime import date, timedelta
from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.utils import timezone
from .services import (
    get_daily_sales, get_period_sales,
    get_sales_by_product, get_sales_by_category,
    get_profit_report, get_purchase_report,
    get_inventory_report, get_stock_movement_report,
    get_supplier_report, get_customer_report,
    get_expense_report, get_tax_report,
    get_employee_sales_report, get_product_movement_report,
)
from .renderers import (
    generate_pdf, generate_excel,
    pdf_response, excel_response, data_to_rows,
)


def parse_dates(request):
    start = request.query_params.get("start_date")
    end = request.query_params.get("end_date")
    if start:
        start = date.fromisoformat(start)
    if end:
        end = date.fromisoformat(end)
    return start, end


class ReportViewSet(viewsets.ViewSet):
    permission_classes = (permissions.IsAuthenticated,)

    @action(detail=False, methods=["get"])
    def daily_sales(self, request):
        date_str = request.query_params.get("date")
        d = date.fromisoformat(date_str) if date_str else timezone.now().date()
        data = get_daily_sales(d)
        fmt = request.query_params.get("format", "json")
        if fmt == "pdf":
            rows = [[s["invoice"], s["customer"], s["time"], s["total"], s["paid"], s["by"]] for s in data["sales"]]
            buf = generate_pdf(f"Daily Sales - {d}", ["Invoice", "Customer", "Time", "Total", "Paid", "Cashier"], rows)
            return pdf_response(buf, f"daily_sales_{d}.pdf")
        if fmt == "xlsx":
            rows = [[s["invoice"], s["customer"], s["time"], s["total"], s["paid"], s["by"]] for s in data["sales"]]
            buf = generate_excel(f"Daily Sales {d}", ["Invoice", "Customer", "Time", "Total", "Paid", "Cashier"], rows)
            return excel_response(buf, f"daily_sales_{d}.xlsx")
        return Response({"success": True, "data": data})

    @action(detail=False, methods=["get"])
    def monthly_sales(self, request):
        today = timezone.now().date()
        start = today.replace(day=1)
        end = today
        data = get_period_sales(start, end, "monthly")
        fmt = request.query_params.get("format", "json")
        headers = ["Date", "Total"]
        rows = [[k, v] for k, v in data["daily_breakdown"].items()]
        if fmt == "pdf":
            buf = generate_pdf(f"Monthly Sales - {today.strftime('%B %Y')}", headers, rows)
            return pdf_response(buf, f"monthly_sales_{today.strftime('%Y_%m')}.pdf")
        if fmt == "xlsx":
            buf = generate_excel(f"Monthly Sales {today.strftime('%B %Y')}", headers, rows)
            return excel_response(buf, f"monthly_sales_{today.strftime('%Y_%m')}.xlsx")
        return Response({"success": True, "data": data})

    @action(detail=False, methods=["get"])
    def yearly_sales(self, request):
        today = timezone.now().date()
        start = today.replace(month=1, day=1)
        end = today
        data = get_period_sales(start, end, "yearly")
        fmt = request.query_params.get("format", "json")
        headers = ["Date", "Total"]
        rows = [[k, v] for k, v in data["daily_breakdown"].items()]
        if fmt == "pdf":
            buf = generate_pdf(f"Yearly Sales - {today.year}", headers, rows)
            return pdf_response(buf, f"yearly_sales_{today.year}.pdf")
        if fmt == "xlsx":
            buf = generate_excel(f"Yearly Sales {today.year}", headers, rows)
            return excel_response(buf, f"yearly_sales_{today.year}.xlsx")
        return Response({"success": True, "data": data})

    @action(detail=False, methods=["get"])
    def sales_by_product(self, request):
        start, end = parse_dates(request)
        data = get_sales_by_product(start, end)
        fmt = request.query_params.get("format", "json")
        headers = ["Product", "SKU", "Qty Sold", "Revenue", "Transactions"]
        rows = [[d["product__name"], d["product__sku"], d["total_qty"], str(d["total_revenue"]), d["sale_count"]] for d in data]
        if fmt == "pdf":
            buf = generate_pdf("Sales by Product", headers, rows)
            return pdf_response(buf, "sales_by_product.pdf")
        if fmt == "xlsx":
            buf = generate_excel("Sales by Product", headers, rows)
            return excel_response(buf, "sales_by_product.xlsx")
        return Response({"success": True, "data": data})

    @action(detail=False, methods=["get"])
    def sales_by_category(self, request):
        start, end = parse_dates(request)
        data = get_sales_by_category(start, end)
        fmt = request.query_params.get("format", "json")
        headers = ["Category", "Qty Sold", "Revenue"]
        rows = [[d["product__category__name"], d["total_qty"], str(d["total_revenue"])] for d in data]
        if fmt == "pdf":
            buf = generate_pdf("Sales by Category", headers, rows)
            return pdf_response(buf, "sales_by_category.pdf")
        if fmt == "xlsx":
            buf = generate_excel("Sales by Category", headers, rows)
            return excel_response(buf, "sales_by_category.xlsx")
        return Response({"success": True, "data": data})

    @action(detail=False, methods=["get"])
    def profit(self, request):
        start, end = parse_dates(request)
        data = get_profit_report(start, end)
        return Response({"success": True, "data": data})

    @action(detail=False, methods=["get"])
    def purchases(self, request):
        start, end = parse_dates(request)
        data = get_purchase_report(start, end)
        fmt = request.query_params.get("format", "json")
        headers = ["Order #", "Supplier", "Date", "Status", "Total", "Paid"]
        rows = [[o["order_number"], o["supplier"], o["date"], o["status"], o["total"], o["paid"]] for o in data["orders"]]
        if fmt == "pdf":
            buf = generate_pdf("Purchase Report", headers, rows)
            return pdf_response(buf, "purchase_report.pdf")
        if fmt == "xlsx":
            buf = generate_excel("Purchase Report", headers, rows)
            return excel_response(buf, "purchase_report.xlsx")
        return Response({"success": True, "data": data})

    @action(detail=False, methods=["get"])
    def inventory(self, request):
        data = get_inventory_report()
        fmt = request.query_params.get("format", "json")
        headers = ["SKU", "Name", "Category", "Qty", "Min", "Reorder", "Cost", "Value", "Status"]
        rows = [[p["sku"], p["name"], p["category"], p["quantity"], p["min_stock"], p["reorder_level"], p["cost_price"], p["stock_value"], p["status"]] for p in data["products"]]
        if fmt == "pdf":
            buf = generate_pdf("Inventory Report", headers, rows)
            return pdf_response(buf, "inventory_report.pdf")
        if fmt == "xlsx":
            buf = generate_excel("Inventory Report", headers, rows)
            return excel_response(buf, "inventory_report.xlsx")
        return Response({"success": True, "data": data})

    @action(detail=False, methods=["get"])
    def stock_movement(self, request):
        start, end = parse_dates(request)
        data = get_stock_movement_report(start, end)
        fmt = request.query_params.get("format", "json")
        headers = ["Date", "Product", "SKU", "Type", "Change", "Balance", "Reference", "By"]
        rows = [[m["date"], m["product"], m["sku"], m["type"], m["qty_change"], m["balance_after"], m["reference"], m["by"]] for m in data["movements"]]
        if fmt == "pdf":
            buf = generate_pdf("Stock Movement Report", headers, rows)
            return pdf_response(buf, "stock_movement.pdf")
        if fmt == "xlsx":
            buf = generate_excel("Stock Movement Report", headers, rows)
            return excel_response(buf, "stock_movement.xlsx")
        return Response({"success": True, "data": data})

    @action(detail=False, methods=["get"])
    def suppliers(self, request):
        data = get_supplier_report()
        fmt = request.query_params.get("format", "json")
        headers = ["Company", "Contact", "Phone", "Balance", "Total Purchases", "Orders"]
        rows = [[s["company"], s["contact"], s["phone"], s["balance"], s["total_purchases"], s["order_count"]] for s in data["suppliers"]]
        if fmt == "pdf":
            buf = generate_pdf("Supplier Report", headers, rows)
            return pdf_response(buf, "supplier_report.pdf")
        if fmt == "xlsx":
            buf = generate_excel("Supplier Report", headers, rows)
            return excel_response(buf, "supplier_report.xlsx")
        return Response({"success": True, "data": data})

    @action(detail=False, methods=["get"])
    def customers(self, request):
        data = get_customer_report()
        fmt = request.query_params.get("format", "json")
        headers = ["Name", "Phone", "Email", "Credit Limit", "Outstanding", "Total Sales", "Orders", "Points"]
        rows = [[c["name"], c["phone"], c["email"], c["credit_limit"], c["outstanding"], c["total_sales"], c["sale_count"], c["loyalty_points"]] for c in data["customers"]]
        if fmt == "pdf":
            buf = generate_pdf("Customer Report", headers, rows)
            return pdf_response(buf, "customer_report.pdf")
        if fmt == "xlsx":
            buf = generate_excel("Customer Report", headers, rows)
            return excel_response(buf, "customer_report.xlsx")
        return Response({"success": True, "data": data})

    @action(detail=False, methods=["get"])
    def expenses(self, request):
        start, end = parse_dates(request)
        data = get_expense_report(start, end)
        fmt = request.query_params.get("format", "json")
        headers = ["Category", "Amount", "Date", "Description", "Recorded By"]
        rows = [[e["category"], e["amount"], e["date"], e["description"], e["by"]] for e in data["items"]]
        if fmt == "pdf":
            buf = generate_pdf("Expense Report", headers, rows)
            return pdf_response(buf, "expense_report.pdf")
        if fmt == "xlsx":
            buf = generate_excel("Expense Report", headers, rows)
            return excel_response(buf, "expense_report.xlsx")
        return Response({"success": True, "data": data})

    @action(detail=False, methods=["get"])
    def cash_flow(self, request):
        start, end = parse_dates(request)
        from apps.accounting.services import get_cash_flow as gcf
        data = {"cash_flow": str(gcf(start, end))}
        return Response({"success": True, "data": data})

    @action(detail=False, methods=["get"])
    def tax(self, request):
        start, end = parse_dates(request)
        data = get_tax_report(start, end)
        return Response({"success": True, "data": data})

    @action(detail=False, methods=["get"])
    def employee_sales(self, request):
        start, end = parse_dates(request)
        data = get_employee_sales_report(start, end)
        fmt = request.query_params.get("format", "json")
        headers = ["Employee", "Name", "Total Sales", "Transactions"]
        rows = [[e["username"], e["name"], e["total_sales"], e["sale_count"]] for e in data["employees"]]
        if fmt == "pdf":
            buf = generate_pdf("Employee Sales Report", headers, rows)
            return pdf_response(buf, "employee_sales.pdf")
        if fmt == "xlsx":
            buf = generate_excel("Employee Sales Report", headers, rows)
            return excel_response(buf, "employee_sales.xlsx")
        return Response({"success": True, "data": data})

    @action(detail=False, methods=["get"])
    def product_movement(self, request):
        data = get_product_movement_report()
        fmt = request.query_params.get("format", "json")
        if fmt in ("pdf", "xlsx"):
            fm = data["fast_moving"]
            sm = data["slow_moving"]
            ds = data["dead_stock"]
            if fmt == "xlsx":
                from openpyxl import Workbook
                import io
                wb = Workbook()
                for name, items, headers in [
                    ("Fast Moving", fm, ["Product", "SKU", "Sold", "Stock"]),
                    ("Slow Moving", sm, ["Product", "SKU", "Sold", "Stock"]),
                    ("Dead Stock", ds, ["Product", "SKU", "Stock"]),
                ]:
                    ws = wb.create_sheet(title=name[:31])
                    for ci, h in enumerate(headers, 1):
                        ws.cell(row=1, column=ci, value=h)
                    for ri, item in enumerate(items, 2):
                        for ci, key in enumerate(headers, 1):
                            ws.cell(row=ri, column=ci, value=item.get(key.lower().replace(" ", "_"), ""))
                if "Sheet" in wb.sheetnames:
                    del wb["Sheet"]
                buf = io.BytesIO()
                wb.save(buf)
                buf.seek(0)
                return excel_response(buf, "product_movement.xlsx")
            if fmt == "pdf":
                buf = generate_pdf("Fast Moving Products", ["Product", "SKU", "Sold", "Stock"], [[i["name"], i["sku"], i["sold"], i["stock"]] for i in fm])
                return pdf_response(buf, "fast_moving.pdf")
        return Response({"success": True, "data": data})
