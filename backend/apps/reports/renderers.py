import io
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import inch, mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    PageBreak, Image,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse


def generate_pdf(title, headers, rows, filename="report.pdf", landscape_mode=False):
    buffer = io.BytesIO()
    page_size = landscape(A4) if landscape_mode else A4
    doc = SimpleDocTemplate(buffer, pagesize=page_size, title=title)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(title, styles["Title"]))
    elements.append(Spacer(1, 12))

    style_table = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e40af")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f1f5f9")]),
    ])

    data = [headers] + rows
    col_widths = [max(len(str(cell)) * 6 + 10, 50) for cell in headers]
    total_width = sum(col_widths)
    page_width = page_size[0] - 80
    if total_width > page_width:
        col_widths = [w * page_width / total_width for w in col_widths]

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(style_table)
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_excel(title, headers, rows, filename="report.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    header_alignment = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")

    for col_idx in range(1, len(headers) + 1):
        max_length = max(
            len(str(ws.cell(row=row_idx, column=col_idx).value or ""))
            for row_idx in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 4, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def data_to_rows(data, fields):
    rows = []
    for item in data:
        row = []
        for field in fields:
            value = item
            for key in field.split("__"):
                if isinstance(value, dict):
                    value = value.get(key, "")
                else:
                    value = getattr(value, key, "") if hasattr(value, key) else str(value)[:50]
            row.append(str(value) if value is not None else "")
        rows.append(row)
    return rows


def pdf_response(buffer, filename="report.pdf"):
    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def excel_response(buffer, filename="report.xlsx"):
    response = HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
